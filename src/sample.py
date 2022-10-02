import openpyxl.utils
import os

from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import FormulaRule

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = '進捗管理表'


def cellwidth(start, end, wide):
    for length in range(start, end):
        alpha = sheet.cell(row=1, column=length).coordinate[:-1]
        sheet.column_dimensions['{}'.format(alpha)].width = wide


cellwidth(1, 7, 8.38)
cellwidth(2, 3, 17.38)
cellwidth(8, 9, 0.38)
cellwidth(9, 335, 4.88)

for rows in range(1, 100):
    sheet.row_dimensions[rows].height = 19

side1 = Side(style='thin', color='000000')
side2 = Side(style='double', color='000000')
side3 = Side(style='hair', color='000000')

for rows in sheet.iter_rows(min_row=1, min_col=1, max_row=100, max_col=9):
    for cell in rows:
        if cell.row == 5:
            cell.border = Border(right=side1, bottom=side2)
        else:
            cell.border = Border(top=None, right=side1, bottom=None, left=None)

for rows in sheet.iter_rows(min_row=1, min_col=9, max_row=100, max_col=335):
    for cell in rows:
        if cell.row == 5:
            cell.border = Border(right=side3, bottom=side2)
        else:
            cell.border = Border(top=side3, right=side3, bottom=side3, left=side3)

startYear = 2020
startMonth = 2
startDay = 1

sheet["B5"].value = "タスク"
sheet["C4"].value = "着手日"
sheet["C5"].value = "予定"
sheet["D5"].value = "実績"
sheet["E4"].value = "完了日"
sheet["E5"].value = "予定"
sheet["F5"].value = "実績"
sheet["G5"].value = "備考"
sheet["I2"].value = startYear
sheet["J2"].value = "年"
sheet["I3"].value = startMonth
sheet["J3"].value = "月"

sheet["I4"].number_format = "d"
sheet["I4"].value = '=DATE(I2, I3, {})'.format(startDay)
sheet["I5"].number_format = "aaa"
sheet["I5"].value = '=I4'

for rows in sheet.iter_rows(min_row=4, min_col=10, max_row=5, max_col=334):
    for cell in rows:
        if cell.row == 4:
            cell.number_format = 'd'
            cell.value = '={}+1'.format(sheet.cell(row=cell.row, column=cell.column-1).coordinate)
        else:
            cell.number_format = 'aaa'
            cell.value = '={}'.format(sheet.cell(row=cell.row-1, column=cell.column).coordinate)

for rows in sheet.iter_rows(min_row=3, min_col=11, max_row=3, max_col=334):
    for cell in rows:
        cell.number_format = "m"
        cell.value = '=IF(DAY({0})=1, {1},"")'.format(sheet.cell(column=cell.column, row=cell.row + 1).coordinate,
                                                      sheet.cell(column=cell.column, row=cell.row + 1).coordinate)


def colorMake(types, start, end):
    return PatternFill(fill_type=types, start_color=start, end_color=end)


grayfill = colorMake('solid', 'd3d3d3', 'd3d3d3')
goldfill = colorMake('solid', 'ffd700', 'ffd700')
bluefill = colorMake('solid', '1e90ff', '1e90ff')

for rows in sheet.iter_rows(min_row=6, min_col=2, max_row=100, max_col=7):
    for cell in rows:
        cell.number_format = "m/d"
sheet.conditional_formatting.add('C6:G100', FormulaRule(formula=['NOT($F6="")'], stopIfTrue=True, fill=grayfill))

sheet.conditional_formatting.add('I4:ME100', FormulaRule(formula=['OR(WEEKDAY(I$5)=1, WEEKDAY(I$5)=7)'], stopIfTrue=True, fill=grayfill))

sheet.conditional_formatting.add('I6:ME100', FormulaRule(formula=['AND($D6<=I$4, $F6>=I$4)'], stopIfTrue=True, fill=goldfill))
sheet.conditional_formatting.add('I6:ME100', FormulaRule(formula=['AND($C6<=I$4, $E6>=I$4)'], stopIfTrue=True, fill=bluefill))

desktop_path = '../excelfile/'
if os.path.exists(desktop_path) == False:
    os.mkdir(desktop_path)
wb.save(desktop_path + '進捗管理表.xlsx')
