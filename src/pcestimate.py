
# -*- coding: utf-8 -*-
import openpyxl as opx
from openpyxl.styles.borders import Border,Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment
newbook = opx.Workbook()
ws = newbook.active

# sheetName = input("input sheet name : ")
# wsName = input("input work sheet name : ")
sheetName = "sample"
wsName = "sample"

ws.title = wsName

side1 = Side(style='thick',color='000000')
side2 = Side(style='thin',color='000000')
side3 = Side(style='double',color='000000')

font1 = Font(name="HGPｺﾞｼｯｸE",size=16,bold=True)

border1 = Border(bottom=side3)
border2 = Border(top=side1,bottom=side1,right=side2,left=side2)

align = Alignment(horizontal='left',vertical='center')

rowIndex=["CPU","CPUグリス","CPUクーラー","マザーボード","GPU","メモリ","カードリーダ","SSD(OS)","SSD1","SSD2","HDD1","HDD2","光学ドライブ","LAN","PCケース","電源","OS"]

colIndex=["パーツ","名称","スペック","価格"]

ws["A1"].value = "BTOショップ"
ws["A2"].value = "BTOパソコン"


for i in range(1,30):
    for j in range(1,30):
        cells=ws.cell(row=i,column=j).coordinate
        ws["{}".format(cells)].alignment=align

        if j == 1 and i != 3:
            ws["{}".format(cells)].border=border1
            ws["{}".format(cells)].font = font1

        if i == 4:
            ws["{}".format(cells)].border=border2
            ws["{}".format(cells)].font=font1

for row in range(len(rowIndex)+10):
    ws.row_dimensions[row].height=30

for cols in range(len(colIndex)+5):
        alpha = ws.cell(row=1, column=cols+1).coordinate[:-1]
        if cols + 1 >= 2 and cols + 1 <= 3:
            ws.column_dimensions["{}".format(alpha)].width = 70
        else:
            ws.column_dimensions["{}".format(alpha)].width = 20

for i,val in enumerate(rowIndex):
    itm = ws.cell(row=i+5,column=1).coordinate
    ws["{}".format(itm)].value = val
    # ws.cell(row=i+4,column=1).border=border1
    # ws.cell(row=i+4,column=1).font=font1

for i,val in enumerate(colIndex):
    itm = ws.cell(row=4,column=i+1).coordinate
    ws["{}".format(itm)].value = val
    # ws["{}".format(itm)].border=border2
    # ws["{}".format(itm)].font=font1
savedir = '../excelfile/'
newbook.save(savedir+'pcestimate.xlsx')